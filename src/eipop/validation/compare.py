"""Cell-by-cell comparison of populated workbook against reference.

DEVELOPMENT TOOL ONLY. This module exists to validate the calculation
engine and Excel writer against the 270-10-1 reference project. It is
not part of the production pipeline — in production, there is no
reference workbook to compare against; the populated workbook IS the
output.

Two validation modes:
1. Python calc validation: Compare Python-calculated values directly
   against reference workbook values (no Excel recalc needed).
2. Workbook validation: Compare a populated .xlsm (after Excel recalc)
   against the reference (both loaded with data_only=True).
"""

from dataclasses import dataclass
from enum import Enum
from pathlib import Path

import openpyxl

from eipop.excel.column_map import COLUMN_MAP, ColType


class CellStatus(Enum):
    MATCH = "MATCH"
    MISMATCH = "MISMATCH"
    SKIP = "SKIP"  # formula column not compared in this mode
    MISSING = "MISSING"  # expected value exists but actual is None
    EMPTY = "EMPTY"  # both are None


@dataclass
class CellResult:
    column: str
    field: str
    col_type: ColType
    expected: object
    actual: object
    status: CellStatus
    error_pct: float | None = None  # relative error for numeric comparisons


def compare_row_to_reference(
    reference_path: str | Path,
    row: int,
    actual_values: dict[str, object],
    rel_tol: float = 1e-3,
) -> list[CellResult]:
    """Compare Python-calculated values against the reference workbook.

    This validates the calculation engine without needing Excel to
    recalculate formulas. Compares all columns where we have actual
    values — both input columns (should be exact) and calculated
    columns (tests our calc engine).

    Args:
        reference_path: Path to reference .xlsm workbook
        row: Row number to compare (e.g. 7 for fire pump)
        actual_values: Dict of {column_map_field: calculated_value}
        rel_tol: Relative tolerance for numeric comparisons

    Returns:
        List of CellResult for every mapped column
    """
    reference_path = Path(reference_path)
    wb = openpyxl.load_workbook(reference_path, data_only=True)
    ws = wb["Proc"]

    results = []
    for col_letter, field_name, col_type in COLUMN_MAP:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        expected = ws.cell(row=row, column=col_idx).value

        if field_name not in actual_values:
            results.append(CellResult(
                column=col_letter, field=field_name, col_type=col_type,
                expected=expected, actual=None,
                status=CellStatus.SKIP,
            ))
            continue

        actual = actual_values[field_name]

        if expected is None and actual is None:
            results.append(CellResult(
                column=col_letter, field=field_name, col_type=col_type,
                expected=expected, actual=actual,
                status=CellStatus.EMPTY,
            ))
            continue

        if expected is None and actual is not None:
            results.append(CellResult(
                column=col_letter, field=field_name, col_type=col_type,
                expected=expected, actual=actual,
                status=CellStatus.MISMATCH,
            ))
            continue

        if expected is not None and actual is None:
            results.append(CellResult(
                column=col_letter, field=field_name, col_type=col_type,
                expected=expected, actual=actual,
                status=CellStatus.MISSING,
            ))
            continue

        # Compare
        status, error_pct = _compare_values(expected, actual, rel_tol)
        results.append(CellResult(
            column=col_letter, field=field_name, col_type=col_type,
            expected=expected, actual=actual,
            status=status, error_pct=error_pct,
        ))

    wb.close()
    return results


def compare_workbooks(
    reference_path: str | Path,
    populated_path: str | Path,
    row: int,
    rel_tol: float = 1e-3,
) -> list[CellResult]:
    """Compare a populated workbook against the reference, cell by cell.

    Both workbooks are loaded with data_only=True, which means the
    populated workbook must have been opened and saved in Excel first
    so that formula values are cached.

    Args:
        reference_path: Path to reference .xlsm
        populated_path: Path to populated .xlsm (after Excel recalc)
        row: Row number to compare
        rel_tol: Relative tolerance for numeric comparisons

    Returns:
        List of CellResult for every mapped column
    """
    reference_path = Path(reference_path)
    populated_path = Path(populated_path)

    wb_ref = openpyxl.load_workbook(reference_path, data_only=True)
    wb_pop = openpyxl.load_workbook(populated_path, data_only=True)
    ws_ref = wb_ref["Proc"]
    ws_pop = wb_pop["Proc"]

    results = []
    for col_letter, field_name, col_type in COLUMN_MAP:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        expected = ws_ref.cell(row=row, column=col_idx).value
        actual = ws_pop.cell(row=row, column=col_idx).value

        if expected is None and actual is None:
            status = CellStatus.EMPTY
            error_pct = None
        elif expected is None:
            status = CellStatus.MISMATCH
            error_pct = None
        elif actual is None:
            status = CellStatus.MISSING
            error_pct = None
        else:
            status, error_pct = _compare_values(expected, actual, rel_tol)

        results.append(CellResult(
            column=col_letter, field=field_name, col_type=col_type,
            expected=expected, actual=actual,
            status=status, error_pct=error_pct,
        ))

    wb_ref.close()
    wb_pop.close()
    return results


def _compare_values(
    expected: object, actual: object, rel_tol: float
) -> tuple[CellStatus, float | None]:
    """Compare two values with type-appropriate logic."""
    # Numeric comparison
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        if expected == 0:
            if actual == 0:
                return CellStatus.MATCH, 0.0
            else:
                return CellStatus.MISMATCH, None
        error_pct = abs(actual - expected) / abs(expected)
        if error_pct <= rel_tol:
            return CellStatus.MATCH, error_pct
        else:
            return CellStatus.MISMATCH, error_pct

    # Boolean comparison
    if isinstance(expected, bool) and isinstance(actual, bool):
        if expected == actual:
            return CellStatus.MATCH, None
        return CellStatus.MISMATCH, None

    # String comparison (exact match)
    if str(expected) == str(actual):
        return CellStatus.MATCH, None
    return CellStatus.MISMATCH, None


def format_report(results: list[CellResult], show_all: bool = False) -> str:
    """Format comparison results as a readable report.

    Args:
        results: List of CellResult from a comparison
        show_all: If True, show all columns. If False, only show
                  MATCH/MISMATCH/MISSING (skip SKIP and EMPTY).

    Returns:
        Formatted report string
    """
    lines = []
    match_count = sum(1 for r in results if r.status == CellStatus.MATCH)
    mismatch_count = sum(1 for r in results if r.status == CellStatus.MISMATCH)
    missing_count = sum(1 for r in results if r.status == CellStatus.MISSING)
    skip_count = sum(1 for r in results if r.status == CellStatus.SKIP)
    total_compared = match_count + mismatch_count + missing_count

    lines.append(f"{'='*70}")
    lines.append(f"VALIDATION REPORT")
    lines.append(f"{'='*70}")
    lines.append(f"Compared: {total_compared}  |  Match: {match_count}  |  "
                 f"Mismatch: {mismatch_count}  |  Missing: {missing_count}  |  "
                 f"Skipped: {skip_count}")
    lines.append(f"{'='*70}")

    if mismatch_count > 0 or missing_count > 0:
        lines.append("")
        lines.append("FAILURES:")
        lines.append(f"{'Col':<5} {'Field':<25} {'Expected':<20} {'Actual':<20} {'Error%':<10}")
        lines.append(f"{'-'*5} {'-'*25} {'-'*20} {'-'*20} {'-'*10}")
        for r in results:
            if r.status in (CellStatus.MISMATCH, CellStatus.MISSING):
                exp_str = _fmt_val(r.expected)
                act_str = _fmt_val(r.actual)
                err_str = f"{r.error_pct*100:.4f}%" if r.error_pct is not None else "—"
                lines.append(f"{r.column:<5} {r.field:<25} {exp_str:<20} {act_str:<20} {err_str:<10}")

    if show_all:
        lines.append("")
        lines.append("ALL RESULTS:")
        lines.append(f"{'Col':<5} {'Field':<25} {'Type':<8} {'Status':<10} {'Expected':<20} {'Actual':<20}")
        lines.append(f"{'-'*5} {'-'*25} {'-'*8} {'-'*10} {'-'*20} {'-'*20}")
        for r in results:
            exp_str = _fmt_val(r.expected)
            act_str = _fmt_val(r.actual)
            lines.append(
                f"{r.column:<5} {r.field:<25} {r.col_type.value:<8} "
                f"{r.status.value:<10} {exp_str:<20} {act_str:<20}"
            )

    if mismatch_count == 0 and missing_count == 0:
        lines.append("")
        lines.append("ALL CHECKS PASSED")

    return "\n".join(lines)


def _fmt_val(val: object) -> str:
    """Format a value for display."""
    if val is None:
        return "None"
    if isinstance(val, float):
        if abs(val) < 0.001 and val != 0:
            return f"{val:.6e}"
        return f"{val:.6f}"
    if isinstance(val, str) and len(val) > 18:
        return val[:15] + "..."
    return str(val)
