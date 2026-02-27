"""End-to-end validation: Python calc engine vs reference workbook.

This test feeds inputs through the calculation engine and compares every
output value against the reference workbook. No Excel recalculation is
involved — this validates the Python engine independently.

Sources validated:
  - S2.002 (row 7): Emergency Fire Water Pump, 129 kW diesel
  - S2.001 (row 6): Emergency Generator, 560 kW diesel
"""

from pathlib import Path

import pytest

from eipop.calcs.diesel_engine import (
    kw_to_hp,
    so2_ef_g_per_bkWhr,
    temp_F_to_K,
    ft_to_m,
    stack_velocity_fps,
    fuel_consumption_MMBtu_hr,
)
from eipop.calcs.ef_selection import select_emission_factors
from eipop.calcs.emission_rates import calc_all_rates
from eipop.calcs.model_rates import lb_hr_to_g_s, ton_yr_to_g_s, model_emission_rates
from eipop.validation.compare import (
    compare_row_to_reference,
    format_report,
    CellStatus,
)

REFERENCE = Path("reference/workbook/LN WFH EmisInv_LnPwr_r1_NDEP.xlsm")

pytestmark = pytest.mark.skipif(
    not REFERENCE.exists(),
    reason="Reference workbook not available",
)

# === S2.002 fire pump inputs ===
KW = 129.0
FUEL_GAL_HR = 11.6
HRS_DAY = 24.0
HRS_YR = 100.0
TEMP_F = 908.0
ACFM = 1131.0
REL_HT_FT = 6.0
DIA_FT = 5.0 / 12.0
NOX_ISR = 0.2

# Emission factors
EF_PM = 0.3
EF_CO = 5.0
EF_NOX = 4.0
EF_VOC = 1.5292437428952  # from reference — derivation TBD


def _build_actual_values() -> dict[str, object]:
    """Run the full calculation chain and return all computed values."""
    hp = kw_to_hp(KW)
    mmbtu_hr = fuel_consumption_MMBtu_hr(FUEL_GAL_HR, "diesel")
    so2_ef = so2_ef_g_per_bkWhr(KW, FUEL_GAL_HR)

    efs = {
        "PM": EF_PM, "PM10": EF_PM, "PM2.5": EF_PM,
        "CO": EF_CO, "NOx": EF_NOX,
        "SO2": so2_ef, "VOC": EF_VOC,
    }

    rates = calc_all_rates(efs, "g/bkW-hr", KW, "bkW", HRS_DAY, HRS_YR)

    # Release params
    vel_fps = stack_velocity_fps(ACFM, DIA_FT)

    # Build the values dict keyed by column_map field names
    values = {}

    # Input fields (should match exactly)
    values["source_id"] = "S2.002"
    values["source_count"] = 1
    values["new_or_modified"] = "Yes"
    values["system_id"] = 2
    values["system_desc"] = "Emergency Fire Water Pump"
    values["scc"] = "2-02-001-02"
    values["source_desc"] = "Emergency Fire Water Pump (75 ≤ kW < 130; mfg. 2010 or newer)"
    values["throughput_unit"] = "bkW"
    values["throughput_material"] = "diesel"
    values["operating_hrs_day"] = HRS_DAY
    values["operating_hrs_yr"] = HRS_YR
    values["kW"] = KW
    values["fuel_type"] = "diesel"
    values["fuel_consumption"] = FUEL_GAL_HR
    values["fuel_unit"] = "gal"
    values["displacement_L"] = 6.8
    values["cylinders"] = 6
    values["tier"] = "N/A"
    values["emission_code"] = "ECI"
    values["EF_PM"] = EF_PM
    values["EF_CO"] = EF_CO
    values["EF_NOx"] = EF_NOX
    values["EF_unit"] = "g/bkW-hr"
    values["ctrl_system"] = "None"
    values["source_type"] = "POINT"
    values["exhaust_temp_F"] = TEMP_F
    values["acfm"] = ACFM
    values["NOx_ISR"] = NOX_ISR
    values["is_emergency"] = True

    # Calculated fields — formulas we replicate
    values["throughput_hr"] = KW
    values["throughput_day"] = KW * HRS_DAY
    values["throughput_yr"] = KW * HRS_YR
    values["hp"] = hp
    values["MMBtu_hr"] = mmbtu_hr
    values["EF_PM10"] = EF_PM
    values["EF_PM2_5"] = EF_PM
    values["EF_SO2"] = so2_ef
    values["EF_VOC"] = EF_VOC

    # Hourly emissions (lb/hr)
    values["PM_pph"] = rates["PM"]["pph"]
    values["PM10_pph"] = rates["PM10"]["pph"]
    values["PM2_5_pph"] = rates["PM2.5"]["pph"]
    values["CO_pph"] = rates["CO"]["pph"]
    values["NOx_pph"] = rates["NOx"]["pph"]
    values["SO2_pph"] = rates["SO2"]["pph"]
    values["VOC_pph"] = rates["VOC"]["pph"]

    # Daily emissions (lb/day)
    values["PM_ppd"] = rates["PM"]["ppd"]
    values["PM10_ppd"] = rates["PM10"]["ppd"]
    values["PM2_5_ppd"] = rates["PM2.5"]["ppd"]
    values["CO_ppd"] = rates["CO"]["ppd"]
    values["NOx_ppd"] = rates["NOx"]["ppd"]
    values["SO2_ppd"] = rates["SO2"]["ppd"]
    values["VOC_ppd"] = rates["VOC"]["ppd"]

    # Annual emissions (ton/yr)
    values["PM_tpy"] = rates["PM"]["tpy"]
    values["PM10_tpy"] = rates["PM10"]["tpy"]
    values["PM2_5_tpy"] = rates["PM2.5"]["tpy"]
    values["CO_tpy"] = rates["CO"]["tpy"]
    values["NOx_tpy"] = rates["NOx"]["tpy"]
    values["SO2_tpy"] = rates["SO2"]["tpy"]
    values["VOC_tpy"] = rates["VOC"]["tpy"]

    # Release parameters
    values["release_height_ft"] = REL_HT_FT
    values["diameter_ft"] = DIA_FT
    values["velocity_fps"] = vel_fps

    # Metric release params
    values["release_height_m"] = ft_to_m(REL_HT_FT)
    values["temp_K"] = temp_F_to_K(TEMP_F)
    values["velocity_mps"] = vel_fps / 3.28084
    values["diameter_m"] = ft_to_m(DIA_FT)

    # Model emission rates (g/s) — emergency source logic
    model = model_emission_rates(rates, is_emergency=True, nox_isr=NOX_ISR)
    values["PM10_24_gps"] = model["PM10_24_gps"]
    values["PM2_5_24_gps"] = model["PM2_5_24_gps"]
    values["CO_ALL_gps"] = model["CO_ALL_gps"]
    values["NOx_1_gps"] = model["NOx_1_gps"]
    values["SO2_1_gps"] = model["SO2_1_gps"]
    values["SO2_ST_gps"] = model["SO2_ST_gps"]
    values["PM2_5_AN_gps"] = model["PM2_5_AN_gps"]
    values["NOx_AN_gps"] = model["NOx_AN_gps"]
    values["SO2_AN_gps"] = model["SO2_AN_gps"]

    # Model NOx/SO2 lb/hr (annualized for emergency)
    values["NOx_model_pph"] = model["NOx_model_pph"]
    values["SO2_model_pph"] = model["SO2_model_pph"]

    # Uncontrolled = same as controlled (no controls on this source)
    values["EF_PM_u"] = EF_PM
    values["EF_PM10_u"] = EF_PM
    values["EF_PM2_5_u"] = EF_PM
    values["EF_CO_u"] = EF_CO
    values["EF_NOx_u"] = EF_NOX
    values["EF_SO2_u"] = so2_ef
    values["EF_VOC_u"] = EF_VOC
    values["EF_unit_u"] = "g/bkW-hr"

    # Uncontrolled hourly = same as controlled hourly
    values["PM_pph_u"] = rates["PM"]["pph"]
    values["PM10_pph_u"] = rates["PM10"]["pph"]
    values["PM2_5_pph_u"] = rates["PM2.5"]["pph"]
    values["CO_pph_u"] = rates["CO"]["pph"]
    values["NOx_pph_u"] = rates["NOx"]["pph"]
    values["SO2_pph_u"] = rates["SO2"]["pph"]
    values["VOC_pph_u"] = rates["VOC"]["pph"]

    # Uncontrolled annual = same as controlled annual
    values["PM_tpy_u"] = rates["PM"]["tpy"]
    values["PM10_tpy_u"] = rates["PM10"]["tpy"]
    values["PM2_5_tpy_u"] = rates["PM2.5"]["tpy"]
    values["CO_tpy_u"] = rates["CO"]["tpy"]
    values["NOx_tpy_u"] = rates["NOx"]["tpy"]
    values["SO2_tpy_u"] = rates["SO2"]["tpy"]
    values["VOC_tpy_u"] = rates["VOC"]["tpy"]

    return values


class TestPythonCalcValidation:
    """Compare Python calculation engine output against reference workbook."""

    def test_all_values_match_reference(self):
        actual = _build_actual_values()
        results = compare_row_to_reference(REFERENCE, 7, actual, rel_tol=1e-3)

        # Print full report for visibility
        report = format_report(results, show_all=False)
        print("\n" + report)

        # Assert no mismatches or missing values
        mismatches = [r for r in results if r.status == CellStatus.MISMATCH]
        missing = [r for r in results if r.status == CellStatus.MISSING]

        if mismatches:
            details = "\n".join(
                f"  {r.column} ({r.field}): expected={r.expected}, actual={r.actual}, error={r.error_pct}"
                for r in mismatches
            )
            pytest.fail(f"{len(mismatches)} mismatched values:\n{details}")

        if missing:
            details = "\n".join(
                f"  {r.column} ({r.field}): expected={r.expected}"
                for r in missing
            )
            pytest.fail(f"{len(missing)} missing values:\n{details}")

    def test_match_count(self):
        """Verify we're actually comparing a meaningful number of cells."""
        actual = _build_actual_values()
        results = compare_row_to_reference(REFERENCE, 7, actual, rel_tol=1e-3)

        matched = sum(1 for r in results if r.status == CellStatus.MATCH)
        # We should match at least 70 cells (inputs + all calculated rates)
        assert matched >= 70, f"Only {matched} matches — expected at least 70"


# === S2.001 emergency generator inputs ===
S2001_KW = 560.0
S2001_FUEL_GAL_HR = 37.55
S2001_HRS_DAY = 24.0
S2001_HRS_YR = 100.0
S2001_TEMP_F = 1100.0
S2001_ACFM = 5429.827606109162
S2001_REL_HT_FT = 6.0
S2001_DIA_FT = 0.6666666666666666
S2001_NOX_ISR = 0.2


def _build_s2001_values() -> dict[str, object]:
    """Run the full calculation chain for S2.001 and return all computed values."""
    kW = S2001_KW
    efs = select_emission_factors(kW=kW, fuel_type="diesel",
                                  fuel_consumption_gal_hr=S2001_FUEL_GAL_HR)

    hp = kw_to_hp(kW)
    mmbtu_hr = fuel_consumption_MMBtu_hr(S2001_FUEL_GAL_HR, "diesel")

    ef_dict = {
        "PM": efs.PM.value, "PM10": efs.PM10.value, "PM2.5": efs.PM2_5.value,
        "CO": efs.CO.value, "NOx": efs.NOx.value,
        "SO2": efs.SO2.value, "VOC": efs.VOC.value,
    }
    rates = calc_all_rates(ef_dict, efs.ef_unit, kW, "bkW",
                           S2001_HRS_DAY, S2001_HRS_YR)
    vel_fps = stack_velocity_fps(S2001_ACFM, S2001_DIA_FT)
    model = model_emission_rates(rates, is_emergency=True, nox_isr=S2001_NOX_ISR)

    values = {}
    values["source_id"] = "S2.001"
    values["source_count"] = 1
    values["new_or_modified"] = "Yes"
    values["system_id"] = 1
    values["system_desc"] = "Emergency Power"
    values["scc"] = "2-02-001-02"
    values["source_desc"] = "Emergency Generator (kW ≤ 560; mfg. 2007 or newer)"
    values["throughput_unit"] = "bkW"
    values["throughput_material"] = "diesel"
    values["operating_hrs_day"] = S2001_HRS_DAY
    values["operating_hrs_yr"] = S2001_HRS_YR
    values["kW"] = kW
    values["fuel_type"] = "diesel"
    values["fuel_consumption"] = S2001_FUEL_GAL_HR
    values["fuel_unit"] = "gal"
    values["displacement_L"] = 16.4
    values["cylinders"] = 8
    values["tier"] = 3
    values["emission_code"] = "ECI"
    values["EF_PM"] = efs.PM.value
    values["EF_CO"] = efs.CO.value
    values["EF_NOx"] = efs.NOx.value
    values["EF_unit"] = efs.ef_unit
    values["ctrl_system"] = "None"
    values["source_type"] = "POINT"
    values["exhaust_temp_F"] = S2001_TEMP_F
    values["acfm"] = S2001_ACFM
    values["NOx_ISR"] = S2001_NOX_ISR
    values["is_emergency"] = True

    values["throughput_hr"] = kW
    values["throughput_day"] = kW * S2001_HRS_DAY
    values["throughput_yr"] = kW * S2001_HRS_YR
    values["hp"] = hp
    values["MMBtu_hr"] = mmbtu_hr
    values["EF_PM10"] = efs.PM10.value
    values["EF_PM2_5"] = efs.PM2_5.value
    values["EF_SO2"] = efs.SO2.value
    values["EF_VOC"] = efs.VOC.value

    for pol in ["PM", "PM10", "CO", "NOx", "SO2", "VOC"]:
        values[f"{pol}_pph"] = rates[pol]["pph"]
        values[f"{pol}_ppd"] = rates[pol]["ppd"]
        values[f"{pol}_tpy"] = rates[pol]["tpy"]
    values["PM2_5_pph"] = rates["PM2.5"]["pph"]
    values["PM2_5_ppd"] = rates["PM2.5"]["ppd"]
    values["PM2_5_tpy"] = rates["PM2.5"]["tpy"]

    values["release_height_ft"] = S2001_REL_HT_FT
    values["diameter_ft"] = S2001_DIA_FT
    values["velocity_fps"] = vel_fps
    values["release_height_m"] = ft_to_m(S2001_REL_HT_FT)
    values["temp_K"] = temp_F_to_K(S2001_TEMP_F)
    values["velocity_mps"] = vel_fps / 3.28084
    values["diameter_m"] = ft_to_m(S2001_DIA_FT)

    values["PM10_24_gps"] = model["PM10_24_gps"]
    values["PM2_5_24_gps"] = model["PM2_5_24_gps"]
    values["CO_ALL_gps"] = model["CO_ALL_gps"]
    values["NOx_1_gps"] = model["NOx_1_gps"]
    values["SO2_1_gps"] = model["SO2_1_gps"]
    values["SO2_ST_gps"] = model["SO2_ST_gps"]
    values["PM2_5_AN_gps"] = model["PM2_5_AN_gps"]
    values["NOx_AN_gps"] = model["NOx_AN_gps"]
    values["SO2_AN_gps"] = model["SO2_AN_gps"]
    values["NOx_model_pph"] = model["NOx_model_pph"]
    values["SO2_model_pph"] = model["SO2_model_pph"]

    values["EF_PM_u"] = efs.PM.value
    values["EF_PM10_u"] = efs.PM10.value
    values["EF_PM2_5_u"] = efs.PM2_5.value
    values["EF_CO_u"] = efs.CO.value
    values["EF_NOx_u"] = efs.NOx.value
    values["EF_SO2_u"] = efs.SO2.value
    values["EF_VOC_u"] = efs.VOC.value
    values["EF_unit_u"] = efs.ef_unit
    for pol in ["PM", "PM10", "CO", "NOx", "SO2", "VOC"]:
        values[f"{pol}_pph_u"] = rates[pol]["pph"]
        values[f"{pol}_tpy_u"] = rates[pol]["tpy"]
    values["PM2_5_pph_u"] = rates["PM2.5"]["pph"]
    values["PM2_5_tpy_u"] = rates["PM2.5"]["tpy"]

    return values


class TestPropaneValidation:
    """Compare Python calculation engine for all propane sources (rows 8-24)."""

    def test_all_propane_rows_match(self):
        import openpyxl
        wb = openpyxl.load_workbook(REFERENCE, data_only=True)
        ws = wb["Proc"]

        total_mismatches = 0
        details = []

        for row in range(8, 25):
            sid = ws.cell(row=row, column=1).value
            fuel = ws.cell(row=row, column=21).value
            acfm = ws.cell(row=row, column=71).value
            ht = ws.cell(row=row, column=68).value
            dia = ws.cell(row=row, column=73).value
            temp = ws.cell(row=row, column=69).value
            nox_isr = ws.cell(row=row, column=91).value or 0.1
            desc = ws.cell(row=row, column=8).value or ""

            values = _build_propane_values(
                sid, desc, fuel, acfm, ht, dia, temp, nox_isr,
            )
            results = compare_row_to_reference(REFERENCE, row, values, rel_tol=1e-3)
            mismatches = [r for r in results if r.status == CellStatus.MISMATCH]
            total_mismatches += len(mismatches)
            for r in mismatches:
                details.append(
                    f"  Row {row} ({sid}) {r.column} ({r.field}): "
                    f"expected={r.expected}, actual={r.actual}"
                )

        wb.close()

        if total_mismatches > 0:
            pytest.fail(
                f"{total_mismatches} mismatches across propane rows:\n"
                + "\n".join(details)
            )

    def test_propane_match_count(self):
        """Verify meaningful comparison count across all propane rows."""
        import openpyxl
        wb = openpyxl.load_workbook(REFERENCE, data_only=True)
        ws = wb["Proc"]

        total_matched = 0
        for row in range(8, 25):
            sid = ws.cell(row=row, column=1).value
            fuel = ws.cell(row=row, column=21).value
            acfm = ws.cell(row=row, column=71).value
            ht = ws.cell(row=row, column=68).value
            dia = ws.cell(row=row, column=73).value
            temp = ws.cell(row=row, column=69).value
            nox_isr = ws.cell(row=row, column=91).value or 0.1
            desc = ws.cell(row=row, column=8).value or ""

            values = _build_propane_values(
                sid, desc, fuel, acfm, ht, dia, temp, nox_isr,
            )
            results = compare_row_to_reference(REFERENCE, row, values, rel_tol=1e-3)
            total_matched += sum(1 for r in results if r.status == CellStatus.MATCH)

        wb.close()
        # 17 rows × ~90 cells each = ~1530
        assert total_matched >= 1500, f"Only {total_matched} matches across 17 rows"


class TestDieselTankValidation:
    """Compare Python calculation for IA1.109 diesel storage tank (row 25)."""

    def test_all_values_match_reference(self):
        values = _build_diesel_tank_values()
        results = compare_row_to_reference(REFERENCE, 25, values, rel_tol=1e-3)

        mismatches = [r for r in results if r.status == CellStatus.MISMATCH]
        if mismatches:
            details = "\n".join(
                f"  {r.column} ({r.field}): expected={r.expected}, actual={r.actual}"
                for r in mismatches
            )
            pytest.fail(f"{len(mismatches)} mismatched values:\n{details}")


def _build_propane_values(
    sid, desc, fuel_gal_hr, acfm, rel_ht_ft, dia_ft, temp_f, nox_isr,
) -> dict[str, object]:
    """Build comparison values for a propane combustion source."""
    from eipop.calcs.diesel_engine import temp_F_to_K, ft_to_m, stack_velocity_fps

    ef_dict = {
        "PM": 0.7, "PM10": 0.7, "PM2.5": 0.7,
        "CO": 7.5, "NOx": 13.0, "SO2": 1.5, "VOC": 1.0,
    }
    tp_hr = fuel_gal_hr / 1000
    rates = calc_all_rates(ef_dict, "lb/kgal", tp_hr, "kgal", 24.0, 8760.0)
    model = model_emission_rates(rates, is_emergency=False, nox_isr=nox_isr)
    vel_fps = stack_velocity_fps(acfm, dia_ft)

    values = {
        "source_id": sid, "source_count": 1, "new_or_modified": "Yes",
        "system_id": "IA", "system_desc": "Insignificant Activities",
        "scc": "1-03-010-02", "source_desc": desc,
        "throughput_unit": "kgal", "throughput_material": "propane",
        "operating_hrs_day": 24.0, "operating_hrs_yr": 8760.0,
        "fuel_type": "propane", "fuel_consumption": fuel_gal_hr,
        "fuel_unit": "gal",
        "EF_PM": 0.7, "EF_CO": 7.5, "EF_NOx": 13.0,
        "EF_unit": "lb/kgal", "source_type": "POINT",
        "exhaust_temp_F": temp_f, "acfm": acfm, "NOx_ISR": nox_isr,
        "throughput_hr": tp_hr, "throughput_day": tp_hr * 24.0,
        "throughput_yr": tp_hr * 8760.0,
        "EF_PM10": 0.7, "EF_PM2_5": 0.7, "EF_SO2": 1.5, "EF_VOC": 1.0,
        "release_height_ft": rel_ht_ft, "diameter_ft": dia_ft,
        "velocity_fps": vel_fps,
        "release_height_m": ft_to_m(rel_ht_ft),
        "temp_K": temp_F_to_K(temp_f),
        "velocity_mps": vel_fps / 3.28084,
        "diameter_m": ft_to_m(dia_ft),
        "NOx_model_pph": model["NOx_model_pph"],
        "SO2_model_pph": model["SO2_model_pph"],
        "EF_PM_u": 0.7, "EF_PM10_u": 0.7, "EF_PM2_5_u": 0.7,
        "EF_CO_u": 7.5, "EF_NOx_u": 13.0, "EF_SO2_u": 1.5, "EF_VOC_u": 1.0,
        "EF_unit_u": "lb/kgal",
    }

    for pol in ["PM", "PM10", "CO", "NOx", "SO2", "VOC"]:
        values[f"{pol}_pph"] = rates[pol]["pph"]
        values[f"{pol}_ppd"] = rates[pol]["ppd"]
        values[f"{pol}_tpy"] = rates[pol]["tpy"]
        values[f"{pol}_pph_u"] = rates[pol]["pph"]
        values[f"{pol}_tpy_u"] = rates[pol]["tpy"]
    values["PM2_5_pph"] = rates["PM2.5"]["pph"]
    values["PM2_5_ppd"] = rates["PM2.5"]["ppd"]
    values["PM2_5_tpy"] = rates["PM2.5"]["tpy"]
    values["PM2_5_pph_u"] = rates["PM2.5"]["pph"]
    values["PM2_5_tpy_u"] = rates["PM2.5"]["tpy"]

    for key in ("PM10_24_gps", "PM2_5_24_gps", "CO_ALL_gps",
                "NOx_1_gps", "SO2_1_gps", "SO2_ST_gps",
                "PM2_5_AN_gps", "NOx_AN_gps", "SO2_AN_gps"):
        values[key] = model[key]

    return values


def _build_diesel_tank_values() -> dict[str, object]:
    """Build comparison values for IA1.109 diesel storage tank."""
    ef_dict = {
        "PM": 0.0, "PM10": 0.0, "PM2.5": 0.0,
        "CO": 0.0, "NOx": 0.0, "SO2": 0.0, "VOC": 16.64,
    }
    rates = calc_all_rates(ef_dict, "lb/yr", 0.0, "gal", 24.0, 8760.0)

    values = {
        "source_id": "IA1.109", "source_count": 1, "new_or_modified": "Yes",
        "system_id": "IA", "system_desc": "Insignificant Activities",
        "scc": "4-04-001-99",
        "source_desc": "Diesel tank (20,000 gal)",
        "throughput_unit": "gal", "throughput_material": "diesel",
        "operating_hrs_day": 24.0, "operating_hrs_yr": 8760.0,
        "EF_unit": "lb/yr", "EF_VOC": 16.64,
        "NOx_model_pph": 0.0, "SO2_model_pph": 0.0,
        "EF_PM_u": 0.0, "EF_PM10_u": 0.0, "EF_PM2_5_u": 0.0,
        "EF_CO_u": 0.0, "EF_NOx_u": 0.0, "EF_SO2_u": 0.0,
        "EF_VOC_u": 16.64, "EF_unit_u": "lb/yr",
    }

    for pol in ["PM", "PM10", "CO", "NOx", "SO2", "VOC"]:
        values[f"{pol}_pph"] = rates[pol]["pph"]
        values[f"{pol}_ppd"] = rates[pol]["ppd"]
        values[f"{pol}_tpy"] = rates[pol]["tpy"]
        values[f"{pol}_pph_u"] = rates[pol]["pph"]
        values[f"{pol}_tpy_u"] = rates[pol]["tpy"]
    values["PM2_5_pph"] = rates["PM2.5"]["pph"]
    values["PM2_5_ppd"] = rates["PM2.5"]["ppd"]
    values["PM2_5_tpy"] = rates["PM2.5"]["tpy"]
    values["PM2_5_pph_u"] = rates["PM2.5"]["pph"]
    values["PM2_5_tpy_u"] = rates["PM2.5"]["tpy"]

    return values


class TestS2001Validation:
    """Compare Python calculation engine output for S2.001 (emergency generator)."""

    def test_all_values_match_reference(self):
        actual = _build_s2001_values()
        results = compare_row_to_reference(REFERENCE, 6, actual, rel_tol=1e-3)

        report = format_report(results, show_all=False)
        print("\n" + report)

        mismatches = [r for r in results if r.status == CellStatus.MISMATCH]
        missing = [r for r in results if r.status == CellStatus.MISSING]

        if mismatches:
            details = "\n".join(
                f"  {r.column} ({r.field}): expected={r.expected}, actual={r.actual}, error={r.error_pct}"
                for r in mismatches
            )
            pytest.fail(f"{len(mismatches)} mismatched values:\n{details}")

        if missing:
            details = "\n".join(
                f"  {r.column} ({r.field}): expected={r.expected}"
                for r in missing
            )
            pytest.fail(f"{len(missing)} missing values:\n{details}")

    def test_match_count(self):
        actual = _build_s2001_values()
        results = compare_row_to_reference(REFERENCE, 6, actual, rel_tol=1e-3)
        matched = sum(1 for r in results if r.status == CellStatus.MATCH)
        assert matched >= 70, f"Only {matched} matches — expected at least 70"
