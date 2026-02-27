"""Test Excel writer: populate template and verify input cells."""

import os
from pathlib import Path

import openpyxl
import pytest

from eipop.excel.writer import populate_template
from eipop.excel.column_map import COLUMN_MAP, ColType
from eipop.models.source import SourceSpec
from eipop.models.emission_factors import EmissionFactors, PollutantFactor


TEMPLATE = Path("templates/LN_WFH_EmisInv_template.xlsm")
OUTPUT = Path("output/test_writer_output.xlsm")

# Skip if template doesn't exist (CI without proprietary files)
pytestmark = pytest.mark.skipif(
    not TEMPLATE.exists(),
    reason="Template workbook not available",
)


@pytest.fixture
def fire_pump_source():
    return SourceSpec(
        source_id="S2.002",
        source_count=1,
        new_or_modified="Yes",
        system_id=2,
        system_desc="Emergency Fire Water Pump",
        scc="2-02-001-02",
        source_desc="Emergency Fire Water Pump (75 ≤ kW < 130; mfg. 2010 or newer)",
        kW=129.0,
        fuel_type="diesel",
        fuel_consumption=11.6,
        fuel_unit="gal",
        displacement_L=6.8,
        cylinders=6,
        tier="N/A",
        operating_hrs_day=24.0,
        operating_hrs_yr=100.0,
        is_emergency=True,
        throughput_unit="bkW",
        source_type="POINT",
        exhaust_temp_F=908.0,
        exhaust_flow_acfm=1131.0,
        emission_code="ECI",
        equipment_reference="Example: Clarke Fire Pump Engines (JU6H-UFABL0)",
        location_reference="G. Young (Target) 2/10/25; beside maintenance bldg",
        cfr_reference="40 CFR 60, Subpart IIII; 40 CFR 63, Subpart ZZZZ",
        state_regulation="NAC 445B.22033, NAC 445B.22024",
        nox_isr=0.2,
    )


@pytest.fixture
def fire_pump_factors():
    ref = "Table 4 to 40 CFR 60, Subpart IIII, 75≤kW<130; SO2 Mass balance based on 15 ppm S; VOC AP-42, Table 3.3-1 (10/96)"
    return EmissionFactors(
        PM=PollutantFactor(0.3, "40 CFR 60 Table 4", "nsps_tier"),
        PM10=PollutantFactor(0.3, "40 CFR 60 Table 4", "nsps_tier"),
        PM2_5=PollutantFactor(0.3, "40 CFR 60 Table 4", "nsps_tier"),
        CO=PollutantFactor(5.0, "40 CFR 60 Table 4", "nsps_tier"),
        NOx=PollutantFactor(4.0, "40 CFR 60 Table 4", "nsps_tier"),
        SO2=PollutantFactor(0.008678900799931303, "Mass balance 15 ppm S", "mass_balance"),
        VOC=PollutantFactor(1.5292437428952, "AP-42 Table 3.3-1", "ap42"),
        ef_unit="g/bkW-hr",
        reference=ref,
    )


@pytest.fixture(autouse=True)
def setup_output_dir():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    yield
    if OUTPUT.exists():
        OUTPUT.unlink()


class TestPopulateTemplate:
    def test_writes_input_cells(self, fire_pump_source, fire_pump_factors):
        populate_template(TEMPLATE, OUTPUT, 7, fire_pump_source, fire_pump_factors)

        wb = openpyxl.load_workbook(OUTPUT, keep_vba=True)
        ws = wb["Proc"]

        # Verify key input cells
        assert ws["A7"].value == "S2.002"
        assert ws["F7"].value == "Emergency Fire Water Pump"
        assert ws["S7"].value == 129.0
        assert ws["AC7"].value == 0.3
        assert ws["AF7"].value == 5.0
        assert ws["AG7"].value == 4.0
        assert ws["AJ7"].value == "g/bkW-hr"
        assert ws["N7"].value == 24.0
        assert ws["O7"].value == 100.0
        assert ws["BQ7"].value == 908.0

        wb.close()

    def test_preserves_formulas(self, fire_pump_source, fire_pump_factors):
        populate_template(TEMPLATE, OUTPUT, 7, fire_pump_source, fire_pump_factors)

        wb = openpyxl.load_workbook(OUTPUT, keep_vba=True)
        ws = wb["Proc"]

        # Formula columns should still contain formulas, not values
        b7 = ws["B7"].value
        assert isinstance(b7, str) and b7.startswith("="), f"B7 should be formula, got: {b7}"

        i7 = ws["I7"].value
        assert isinstance(i7, str) and i7.startswith("="), f"I7 should be formula, got: {i7}"

        r7 = ws["R7"].value
        assert isinstance(r7, str) and r7.startswith("="), f"R7 should be formula, got: {r7}"

        wb.close()

    def test_preserves_vba(self, fire_pump_source, fire_pump_factors):
        populate_template(TEMPLATE, OUTPUT, 7, fire_pump_source, fire_pump_factors)

        wb = openpyxl.load_workbook(OUTPUT, keep_vba=True)
        assert wb.vba_archive is not None, "VBA macros should be preserved"
        wb.close()

    def test_preserves_other_sheets(self, fire_pump_source, fire_pump_factors):
        populate_template(TEMPLATE, OUTPUT, 7, fire_pump_source, fire_pump_factors)

        wb = openpyxl.load_workbook(OUTPUT, keep_vba=True)
        assert "Conv" in wb.sheetnames
        assert "(40CFR89&1039)" in wb.sheetnames
        assert "B-bpip01" in wb.sheetnames

        # Spot check Conv values are intact
        ws_conv = wb["Conv"]
        assert ws_conv.cell(row=35, column=3).value == 1.341  # hp/kW

        wb.close()

    def test_only_writes_target_row(self, fire_pump_source, fire_pump_factors):
        populate_template(TEMPLATE, OUTPUT, 7, fire_pump_source, fire_pump_factors)

        wb = openpyxl.load_workbook(OUTPUT, keep_vba=True)
        ws = wb["Proc"]

        # Rows 6, 8 should still be empty (cleared in template)
        assert ws["A6"].value is None
        assert ws["A8"].value is None

        wb.close()
